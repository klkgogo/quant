from openpyxl import *
import sys
def not_empty_row(row):
    if row[0].value is None:
        return False
    else:
        return True

def title_row(row):
    if row[0].value == "Item Number":
        return True
    else:
        return False


def merge(target, src):
    for row in src.iter_rows():
        if not title_row(row) and not_empty_row(row):
            target.append((cell.value for cell in row))

def copy_title(target, src):
    for row in src.iter_rows():
        if title_row(row):
            target.append((cell.value for cell in row))
            return


if __name__ == "__main__":
    argv  = ['merge.py', 'new.xlsx', 'C:\\Users\\henrykong\\Desktop\\Navy_副板_Draft_BOM_20180629.xlsx',  'C:\\Users\\henrykong\\Desktop\\Navy_主FPC_Draft_BOM_20180629.xlsx', 'C:\\Users\\henrykong\\Desktop\\Navy_主板_Draft_BOM_20180629.xlsx']
    print(sys.argv)
    if (len(sys.argv) < 3):
        print("usage: merge output.xls input1.xls input2.xls input3.xls ....")
        sys.exit(2)

    target_name = sys.argv[1]
    target_wb = Workbook()
    target_sheet = target_wb.active
    input_files = sys.argv[2:]
    init = False
    for file_name in input_files:
        print("merge ", file_name)
        wb = load_workbook(file_name)
        sheetname = wb.sheetnames[0]
        sheet = wb[sheetname]
        if not init:
            copy_title(target_sheet, sheet)
            init = True
        merge(target_sheet, sheet)
        wb.close()

    colA = target_sheet['A']
    index = 1
    for row in target_sheet.iter_rows():
        if not title_row(row):
            row[0].value = index
            index += 1

    target_wb.save(target_name)
    print("merge end")
    print("save to ", target_name)
